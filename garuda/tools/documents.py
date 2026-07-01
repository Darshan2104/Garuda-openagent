"""Read PDF and spreadsheet files from the workspace."""

import base64
import shlex
from pathlib import Path

from garuda.tools.protocol import ToolContext
from garuda.types import ToolResult
from garuda.workspace.protocol import Environment


async def _read_file_bytes(env: Environment, path: str) -> bytes:
    resolved = path if path.startswith("/") else f"{env.workspace_root.rstrip('/')}/{path}"
    result = await env.execute(f"base64 -w0 {shlex.quote(resolved)} 2>/dev/null || base64 {shlex.quote(resolved)}")
    if result.exit_code != 0:
        raise FileNotFoundError(result.stderr or f"Cannot read binary file: {path}")
    return base64.b64decode(result.stdout.strip())


def _read_pdf_bytes(data: bytes, max_chars: int = 50_000) -> str:
    from io import BytesIO

    from pypdf import PdfReader

    reader = PdfReader(BytesIO(data))
    parts: list[str] = []
    for index, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            parts.append(f"--- Page {index + 1} ---\n{text}")
    content = "\n\n".join(parts)
    if len(content) > max_chars:
        content = content[:max_chars] + "\n...[truncated]"
    return content or "(No extractable text in PDF)"


def _read_xlsx_bytes(data: bytes, max_rows: int = 200) -> str:
    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows: list[str] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index >= max_rows:
                rows.append("...[truncated rows]")
                break
            cells = ["" if cell is None else str(cell) for cell in row]
            if any(cells):
                rows.append("\t".join(cells))
        parts.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows))
    workbook.close()
    return "\n\n".join(parts)


class ReadPdfTool:
    name = "read_pdf"
    description = "Extract text content from a PDF file in the workspace."
    parameters = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Path to the PDF file"},
            "max_chars": {"type": "integer", "description": "Max characters to return", "default": 50000},
        },
        "required": ["path"],
    }

    async def execute(self, arguments: dict, env: Environment, ctx: ToolContext) -> ToolResult:
        path = arguments["path"]
        max_chars = int(arguments.get("max_chars", 50_000))
        try:
            import pypdf  # noqa: F401
        except ImportError:
            return ToolResult(
                tool_call_id="",
                content="read_pdf requires: pip install 'garuda-openagent[docs]'",
                is_error=True,
            )
        data = await _read_file_bytes(env, path)
        content = _read_pdf_bytes(data, max_chars=max_chars)
        return ToolResult(tool_call_id="", content=content)


class ReadSpreadsheetTool:
    name = "read_spreadsheet"
    description = "Read an Excel (.xlsx) or CSV spreadsheet from the workspace as tab-separated text."
    parameters = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Path to .xlsx or .csv file"},
            "max_rows": {"type": "integer", "description": "Max rows per sheet", "default": 200},
        },
        "required": ["path"],
    }

    async def execute(self, arguments: dict, env: Environment, ctx: ToolContext) -> ToolResult:
        path = arguments["path"]
        max_rows = int(arguments.get("max_rows", 200))
        suffix = Path(path).suffix.lower()
        if suffix == ".csv":
            content = await env.read_file(path)
            lines = content.splitlines()[:max_rows]
            return ToolResult(tool_call_id="", content="\n".join(lines))
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            return ToolResult(
                tool_call_id="",
                content="read_spreadsheet requires: pip install 'garuda-openagent[docs]'",
                is_error=True,
            )
        data = await _read_file_bytes(env, path)
        content = _read_xlsx_bytes(data, max_rows=max_rows)
        return ToolResult(tool_call_id="", content=content)
